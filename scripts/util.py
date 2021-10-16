import openpyxl, csv

import urllib, feedparser

from scholarly import scholarly
from scholarly import ProxyGenerator

import xlsxwriter
from bibtexparser.bparser import BibTexParser


def dict_from_string(str):
    str.strip(" \n\t")
    assert str[0] == "@"
    type = str[1 : str.find("{")]
    # For each bibtex key, make it a
    parser = BibTexParser()
    try:
        dict = parser.parse(str).get_entry_dict()
        name = list(dict.keys())[0]
    except Exception as e:
        print("Probably malformed bibtex: \n", str)
        raise e
    dict_ = dict[name]
    dict = {}
    for k in dict_:
        dict[k.lower()] = dict_[k]
    return type, name, dict


def get_arxiv(row):
    if "https://arxiv.org/pdf/" in row[4]:
        id = row[4][22:32]
    elif "https://arxiv.org/ftp/arxiv/papers/" in row[4]:
        id = row[4][40:50]
    else:
        raise ValueError("Invalid arxiv link")
    url = 'http://export.arxiv.org/api/query?id_list={}&start=0&max_results=1'.format(id)
    data = urllib.request.urlopen(url)
    d = feedparser.parse(data.read().decode('utf-8'))
    return d, id


def get_scholarly_result(title):
    search_result = None
    cnt = 1
    while search_result is None:
        try:
            search_result = next(scholarly.search_pubs(title))
            # search_result = scholarly.fill(search_result)
        except:
            print("Generating proxy attempt {}...".format(cnt))
            # Generate proxy to avoid Google banning bots
            try:
                pg = ProxyGenerator()
                pg.FreeProxies()
                scholarly.use_proxy(pg)
                cnt += 1
            except: pass
            search_result = None
    return search_result

def read_spreadsheet(input_fname, input_ext):
    rows_in = []
    if (input_ext == ".xlsx"):
        print("Loading workbook: ", input_fname + input_ext)
        workbook = openpyxl.load_workbook(input_fname + input_ext)
        sheet = workbook.active
        for row in sheet.iter_rows():
            row_in = []
            allNone = True
            for cell in row:
                if cell.value is None:
                    row_in.append("")
                else:
                    allNone = False
                    row_in.append(cell.value)
            if allNone: continue
            rows_in.append(row_in)
    elif (input_ext == ".csv"):
        with open(input_fname + input_ext, newline='', encoding="utf-8") as csvfile:
            reader = csv.reader(csvfile, delimiter=',', quotechar='"')
            for row in reader:
                rows_in.append(row)

    return rows_in


def write_spreadsheet(rows, fname, ext):

    if ext == ".csv":
        with open(fname + ext, "w+", newline='', encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(rows)
    elif ext == ".xlsx":
        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook(fname + ext)
        worksheet = workbook.add_worksheet()
        # Start from the first cell. Rows and columns are zero indexed.
        row = 0
        col = 0

        # Iterate over the data and write it out row by row.
        for r in rows:
            for c in r:
                worksheet.write(row, col, c)
                col += 1
            row += 1
            col = 0
        workbook.close()


def format_bibtex_str(bibtex, cap_keys="ALL", space=True, indent="    ", article_type="article", last_name_first=False):
    """
    Args:
        bibtex: str or dict
        cap_keys: {"ALL", "Initial", "lower"}
        space: boolean, whether to have spcae around "="
    """
    if space:
        equal = " = "
    else:
        equal = "="
    if type(bibtex) is dict:
        name = list(bibtex.keys())[0]
        d = bibtex[name]
        entries = []
        for key in d:
            content = d[key].replace(' \n', ' ').replace('\n', ' ')
            key = key.lower()
            if cap_keys == "ALL":
                key = key.upper()
            elif cap_keys == "Initial":
                key = key[0].upper() + key[1:]
            if key == "AUTHOR":
                if (not last_name_first) and (", " in content):
                    content = " and ".join([" ".join(a.split(", ")[::-1]) for a in content.split(" and ")])
            entries.append(indent + key + equal + "{" + content + "}")
        head = f"@{article_type.lower()}"+"{"+f"{name},\n"
        tail = "\n}"
        texstr = head + ",\n".join(entries) + tail
    else:
        texstr = bibtex
        indents = {"\t", " ", "  ", "    ", "          ", "\r"}
        # print(texstr)
        ind = 0
        cnt = 0
        while ind < len(texstr):
            nl = texstr[ind:].find("\n") + ind
            eq = texstr[nl:].find("=") + nl
            bra = texstr[eq:].find("{") + eq
            if (eq - nl) == -1:
                break

            # Format key
            key = texstr[nl+1:eq].replace(" ", "").replace("\t","").replace("\r","").lower()
            if cap_keys == "ALL":
                key = key.upper()
            elif cap_keys == "Initial":
                key = key[0].upper() + key[1:]

            orig_len = len(texstr)
            texstr = texstr[:nl+1] + key + equal + texstr[bra:]
            bra += len(texstr) - orig_len

            # Skip the value (bracketed)
            left = bra
            if (left - eq) == -1:
                break
            ind = left + 1
            left_cnt = 1
            right_cnt = 0
            while (left_cnt > right_cnt):
                left = texstr[ind:].find("{") + ind
                right = texstr[ind:].find("}") + ind
                # print(left, right, left_cnt, right_cnt, ind, repr(texstr[ind:ind+10]))
                if (left < right) and ((left - ind) != -1):
                    left_cnt += 1
                    ind = left + 1
                elif (right - ind) != -1:
                    right_cnt += 1
                    ind = right + 1
                cnt += 1
                if cnt > 100:
                    raise ValueError("Yiheng wrote a bug... check for infinite while loop here")
            ket = right
            braket = texstr[bra:ket]
            for i in indents:
                braket = braket.replace("\n"+i, " ")
            braket = braket.replace("\n", " ")
            texstr = texstr[:bra] + braket + texstr[ket:]

        if indent is not None:
            for i in indents:
                texstr = texstr.replace("\n"+i+i, "\n")
                texstr = texstr.replace("\n"+i, "\n")

            texstr = texstr.replace("\n", "\n"+indent)
            if texstr[-(2+len(indent)):] == "\n"+indent+"}":
                texstr = texstr[:-(2+len(indent))] + "\n}"

        # Format tailend
        rev_str = texstr[::-1]
        last = rev_str.find("}")
        second_last = rev_str[last+1:].find("}") + last + 1
        texstr = rev_str[second_last:][::-1]+"\n}"
    return texstr


def bibtex_name_from_bibtex(bibtex):
    if type(bibtex) is dict:
        name = list(bibtex.keys())[0]
    else:
        bibtex = bibtex.replace("\r\n", "\n")
        start, end = bibtex.find("{") + 1, bibtex.find(",")
        name = bibtex[start:end].lower()
        return name


def get_authors_from_bibtex(bibtex, last_name_first=False):
    if type(bibtex) is dict:
        authors_str = bibtex['author']
    else:
        author = {"author = {": 10, "author={": 8, "AUTHOR = {": 10, "AUTHOR={": 8}
        for a in author:
            ind = bibtex.find(a)
            if ind >= 0:
                start_ind = ind + author[a]
                end_ind = start_ind + bibtex[start_ind:].find("}")
                authors_str = bibtex[start_ind:end_ind]

    authors_list = authors_str.split(" and ")
    if (not last_name_first) and (", " in authors_str):
        authors_list = [" ".join(a.split(", ")[::-1]) for a in authors_list]
    return authors_list

def get_venue(comment, pub_year):
    conf, year, venue = "", "", ""

    year_range = [str("%02d" % d) for d in range(30)] + [str("20%02d" % d) for d in range(30)]
    for y in year_range:
        if (y in comment):
            y = "20" + y if (len(y) == 2) else y
            if (abs(int(y) - int(pub_year)) < 2):
                year = y
                break

    conf_range = ["CVPR", "SIGGRAPH", "ECCV", "3DV", "NeurIPS", "ICCV", "IROS", "EGSR", "EuroVis", "IJCAI"]
    conf_range += [c.lower() for c in conf_range]
    for c in conf_range:
        if c in comment:
            conf = c
            break

    if conf != "":
        venue = conf + " " + year
    return venue


## Note on citation conventions and generation
"""
Minimally, the keys should be: AUTHOR, TITLE, BOOKTITLE, YEAR,
Our format should have keys: AUTHOR, TITLE, YEAR, MONTH, URL, JOURNAL
"""
KNOWN_FORMATS = ['3DV', 'RSS', 'ARXIV', 'NeurIPS', 'SIGGRAPH', 'GRAPHITE04',
    'ICML', 'ICCV', 'TPAMI', 'CoRL', 'IJCAI', 'ECCV', 'CVPR', 'EGSR', 'ICLR'
]
BIBTEX_INFO = {
    "booktitle": {
        "ARXIV": "ArXiv Pre-print",
        "CVPR": "Proceedings of the IEEE/CVF Conference on Computer Vision and Pattern Recognition (CVPR)",
        "ICCV": "Proceedings of the IEEE International Conference on Computer Vision (ICCV)",
        "ECCV": "Proceedings of the European Conference on Computer Vision (ECCV)",
        "NeurIPS": "Advances in Neural Information Processing Systems (NeurIPS)",
        "3DV": "International Conference on 3D Vision (3DV)",
        "ICML": "International Conference on Machine Learning (ICML)",
        "CoRL": "Proceedings of the Conference on Robot Learning (CoRL)",
        "ICLR": "International Conference on Learning Representations",
        "IJCAI": "Proceedings of the Thirtieth International Joint Conference on Artificial Intelligence (IJCAI)",
        "RSS": "Proceedings of Robotics: Science and Systems",
    },
    "journal": {
        "SIGGRAPH": "ACM Transactions on Graphics (TOG)",
        "TPAMI": "IEEE Transactions on Pattern Analysis and Machine Intelligence",
        "EGSR": "Computer Graphics Forum",
        "ARXIV": "arXiv preprint"
    },
    "month": {
        "CVPR 2020": "Jun",
        "CVPR 2021": "Jun",
        "TPAMI 2021": "Jul",
        "ICCV 2020": "Jun",
        "ICCV 2021": "Oct",
        "CoRL 2020": "Nov",
        "SIGGRAPH 2020": "Nov",
        "SIGGRAPH 2021": "Jul",
        "RSS 2021": "Jul",
        "IJCAI 2021": "Aug",
    },
    "publisher": {
        "SIGGRAPH": "Association for Computing Machinery (ACM)",
        "NeurIPS": "Curran Associates, Inc.",
        "ICML": "PMLR",
        "SIGGRAPH": "Association for Computing Machinery",
        "IJCAI": "International Joint Conferences on Artificial Intelligence Organization",
        "EGSR": "The Eurographics Association and John Wiley & Sons Ltd.",
    },
    "organization": {
        "3DV": "IEEE",
    },
}
TYPE = {
    "CVPR": "inproceedings",
    "ICCV": "inproceedings",
    "ECCV": "inproceedings",
    "NeurIPS": "inproceedings",
    "3DV": "inproceedings",
    "ICML": "inproceedings",
    "ICLR": "inproceedings",
    "RSS": "inproceedings",
    "IJCAI": "inproceedings",
    "CoRL": "inproceedings",
    "SIGGRAPH": "article",
    "TPAMI": "article",
    "EGSR": "article",
    "ARXIV": "article",
}

csv_head = ['Timestamp', 'Email Address', 'Does your work use coordinate(s) as neural network input(s)?', 'Title', 'Nickname (e.g. DeepSDF)', '"Venue (e.g. ""CVPR"".""SIGGRAPH"".or ""ECCV (Oral)"")"', '"Date (earliest release date.e.g. arXiv v1 date)"', '"Year (corresponding to venue e.g. released in 2021.accepted to CVPR 2022.then put ""2022"" for this entry.and ""2021"" for the above)"', 'Bibtex Citation', 'PDF (use ArXiv when possible)', 'Project Webpage (web link)', 'Code Release (Github link)', 'Data Release (link)', '"Talk/Video (link.e.g. youtube)"', 'Supplement PDF (link)', '"Supplement video (link.comma separated if multiple exists)"', 'Keywords', 'Frequency/Positional Encoding', '"Geometry proxy (for non-visual computing papers.choose ""N/A"")"', '"Reconstructs Geometry Only (i.e. no color texture) (for non-visual computing papers.choose ""N/A"")"', 'Training time (hr)', 'Rendering time (FPS)', 'Dataset(s) used (e.g. Tanks and Temples)', '# of input views (e.g. 18 for 18-camera system)', 'Inputs', 'Lighting', 'Authors', 'Bibtex Name', 'UID', 'Citation Count', 'Abstract', 'Coordinates all at once', '"Feature-as-input (coordinate samples feature grid.but coordinate is not supplied as input)"', 'Direct/Indirect Neural Field (one or more dimension built into the network e.g. 2D CNN + z)']
csv_head_key = {'Timestamp': 0, 'Email': 1, 'Screening': 2, 'Title': 3, 'Nickname': 4, 'Venue': 5, 'Date': 6, 'Year': 7, 'Bibtex': 8, 'PDF': 9, 'Website': 10, 'Code': 11, 'Data': 12, 'Talk/Video': 13, 'Supplement PDF': 14, '"Supplement': 15, 'Keywords': 16, 'Frequency/Positional Encoding': 17, 'Geometry proxy': 18, 'Geometry Only': 19, 'Training time': 20, 'Rendering time': 21, 'Dataset(s) used': 22, '# of input views': 23, 'Inputs': 24, 'Lighting': 25, 'Authors': 26, 'Bibtex Name': 27, 'UID': 28, 'Citation Count': 29, 'Abstract': 30, 'Coordinates all at once': 31, 'Feature-as-input': 32, 'Direct/Indirect Neural Field': 33}
# Old sheet
# csv_head_key = {'Timestamp': 0, 'Title': 1, 'Nickname': 2, 'Venue': 23, 'Date': 3, 'Bibtex': 11, 'PDF': 4, 'Authors': 27, 'Bibtex Name': 28, 'Citation Count': 31, 'Abstract': 30, 'UID': 29}
